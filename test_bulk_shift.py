# -*- coding: utf-8 -*-
"""
bulk_shift.py (siniflandirma, payload insasi, Excel ayristirma) ve
/api/bulk-shift/* uc noktalarinin dogrulama davranisini kontrol eder.
Gercek CSM auth/create_ticket cagrisi YAPILMAZ (harici uretim sistemine
gercek istek atmadan test edilebilecek her sey buradadir).
"""

import io
import os
import sys
import tempfile

if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# Temp SQLite DB for the login below -- must be set before `from app import
# app` (app.py reads DATABASE_URL at import time). Never touches the real
# instance/enigma.db.
_tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp_db.close()
os.environ["DATABASE_URL"] = f"sqlite:///{_tmp_db.name}"

from openpyxl import Workbook

from bulk_shift import (
    classify_shift_type,
    build_bulk_ticket_payload,
    parse_excel_rows,
    find_result_columns,
    generate_result_workbook,
    MAX_ROWS,
)
import app as app_module
import logging_utils
from app import app
from dev_test_helpers import login_test_client

# Panelin /api/bulk-shift/upload'ının başarılı bir process_rows() çağrısından
# SONRA yazdığı mail log kaydını gerçek mail_processing_log.jsonl'a değil,
# bu dosyaya yazdırıyoruz (CANLI HATA DÜZELTMESİ: bu kayıt önceden hiç
# yazılmıyordu, panelden yapılan toplu kaydırmalar -- başarılı ya da
# başarısız -- Loglar sayfasında hiç görünmüyordu).
_tmp_mail_log = os.path.join(os.path.dirname(__file__), "_test_bulk_shift_mail_log.jsonl")
if os.path.exists(_tmp_mail_log):
    os.remove(_tmp_mail_log)
logging_utils.LOG_FILE = _tmp_mail_log

passed = 0
failed = 0
results = []


def check(name, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
        results.append(("PASS", name, detail))
    else:
        failed += 1
        results.append(("FAIL", name, detail))


# ==========================================================
# classify_shift_type
# ==========================================================
check("classify: 'Otel Kaynaklı' -> OTEL_KAYNAKLI", classify_shift_type("Otel Kaynaklı") == "OTEL_KAYNAKLI")
check(
    "classify: 'Ödeme Tamamlama' -> ODEME_TAMAMLAMA",
    classify_shift_type("Ödeme Tamamlama") == "ODEME_TAMAMLAMA",
)
check(
    "classify: bilinmeyen metin -> varsayilan OPERASYON_KAYNAKLI",
    classify_shift_type("Bilinmeyen Tip") == "OPERASYON_KAYNAKLI",
)
check("classify: bos string -> varsayilan OPERASYON_KAYNAKLI", classify_shift_type("") == "OPERASYON_KAYNAKLI")

# ==========================================================
# build_bulk_ticket_payload
# ==========================================================
payload = build_bulk_ticket_payload(
    reservation_no="123456",
    shift_type_label="Otel Kaynaklı",
    alternative_text="Toplu Kaydırma İşlemi",
    parent_ticket_uuid="parent-uuid-123",
)
check(
    "payload: parent ticket UUID + relationType ust seviyede dogru yerlestirilmis (LINKED_TICKET)",
    payload["parentTicketUUID"] == "parent-uuid-123" and payload["relationType"] == "LINKED_TICKET",
    {"parentTicketUUID": payload.get("parentTicketUUID"), "relationType": payload.get("relationType")},
)
check(
    "payload: ticketRelationList BOS (parent-child DEGIL, iliskili ticket)",
    payload["ticketRelationList"] == [],
    payload["ticketRelationList"],
)
check(
    "payload: raporlayan sabit 'Onay Kaydırma' kontagi (mail gonderenden degil)",
    payload["partyRole"]["party"]["fullName"] == "Onay Kaydırma"
    and payload["partyRole"]["contactMediumList"][0]["val"] == "onay@tatilbudur.com",
    payload["partyRole"]["party"],
)
check(
    "payload: Otel Kaynakli kirilimi (KAYDIRMA > OTEL_KAYNAKLI) kullanilmis",
    payload["category"]["shortCode"] == "KAYDIRMA" and payload["subCategory"]["shortCode"] == "OTEL_KAYNAKLI",
    {"category": payload["category"], "subCategory": payload["subCategory"]},
)

odeme_payload = build_bulk_ticket_payload(
    reservation_no="999",
    shift_type_label="Ödeme Tamamlama",
    alternative_text="x",
    parent_ticket_uuid="parent-uuid-123",
)
check(
    "payload: Odeme Tamamlama kirilimi (DIGER_ISLEMLER > ODEME_TAMAMLAMA) kullanilmis",
    odeme_payload["category"]["shortCode"] == "DIGER_ISLEMLER"
    and odeme_payload["subCategory"]["shortCode"] == "ODEME_TAMAMLAMA",
    {"category": odeme_payload["category"], "subCategory": odeme_payload["subCategory"]},
)
check(
    "payload: relatedProduct.serviceNumber rezervasyon no'yu tasiyor",
    payload["relatedProduct"]["serviceNumber"] == "123456",
)

# ==========================================================
# build_bulk_ticket_payload -- standalone "ana ticket" (BO/YÇM: bos parent_ticket_uuid)
# ==========================================================
standalone_payload = build_bulk_ticket_payload(
    reservation_no="777",
    shift_type_label="Operasyon Kaynaklı",
    alternative_text="x",
    parent_ticket_uuid="",
)
check(
    "payload (ANA TICKET, kullanici tarafindan istendi): bos parent_ticket_uuid -> parentTicketUUID/relationType alanlari HIC YOK",
    "parentTicketUUID" not in standalone_payload and "relationType" not in standalone_payload,
    standalone_payload,
)
check(
    "payload (ANA TICKET): kirilim/raporlayan iliskili ticket'la AYNI (kullanici onayladi)",
    standalone_payload["category"]["shortCode"] == "KAYDIRMA"
    and standalone_payload["subCategory"]["shortCode"] == "OPERASYON_KAYNAKLI"
    and standalone_payload["partyRole"]["party"]["fullName"] == "Onay Kaydırma",
    standalone_payload,
)

hotel_payload = build_bulk_ticket_payload(
    reservation_no="777",
    shift_type_label="Operasyon Kaynaklı",
    alternative_text="x",
    parent_ticket_uuid="",
    hotel="Örnek Otel",
    room_type="Double",
)
check(
    "payload: Otel/Oda Tipi verilince aciklamaya ekleniyor",
    "Örnek Otel" in hotel_payload["description"] and "Double" in hotel_payload["description"],
    hotel_payload["description"],
)

# ==========================================================
# parse_excel_rows
# ==========================================================
def make_excel(rows, headers=("Rezervasyon No", "Kaydırma Tipi", "Alternatif 1", "parentTicketUUID")):
    wb = Workbook()
    ws = wb.active
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


good_excel = make_excel(
    [
        ("553044193", "Otel Kaynaklı", "Alt 1", "parent-uuid-abc"),
        ("358109758", "Operasyon Kaynaklı", "Alt 2", ""),
        ("", "Boş satır atlanmalı", "x", ""),
    ]
)
parsed = parse_excel_rows(good_excel)
check("parse_excel_rows: 2 gecerli satir donuyor (bos satir atlaniyor)", len(parsed) == 2, parsed)
check("parse_excel_rows: ilk satirin alanlari dogru okunmus", parsed[0]["reservation_no"] == "553044193", parsed[0])
check(
    "parse_excel_rows: bos parentTicketUUID onceki dolu degerden devralinir",
    parsed[1]["parent_ticket_uuid"] == "parent-uuid-abc",
    parsed[1],
)

missing_alt_excel = make_excel([("111", "Otel Kaynaklı", "", "uuid-1")])
parsed2 = parse_excel_rows(missing_alt_excel)
check(
    "parse_excel_rows: bos 'Alternatif 1' icin varsayilan metin kullanilir",
    parsed2[0]["alternative"] == "Toplu Kaydırma İşlemi",
    parsed2,
)

bad_columns_excel = make_excel([("x", "y", "z", "w")], headers=("Yanlis", "Sutunlar", "Burada", "Da"))
try:
    parse_excel_rows(bad_columns_excel)
    check("parse_excel_rows: eksik sutun ValueError firlatir", False, "hata firlatmadi")
except ValueError as e:
    check("parse_excel_rows: eksik sutun ValueError firlatir", "Rezervasyon No" in str(e), str(e))

# ==========================================================
# parse_excel_rows -- BO/YÇM şablonu (parentTicketUUID sütunu HİÇ YOK,
# birden fazla "Alternatif" sütunu, Otel/Oda Tipi opsiyonel sütunları)
# ==========================================================
bo_ycm_excel = make_excel(
    [("346845477", "Operasyon Kaynaklı", "Ege Akdeniz Turu", "5 Gece Otel", "12.07 hareketli")],
    headers=("Rezervasyon No", "Kaydırma Tipi", "Alternatif", "Alternatif", "Alternatif"),
)
bo_parsed = parse_excel_rows(bo_ycm_excel)
check(
    "parse_excel_rows (BO/YÇM, kullanici tarafindan bildirildi): parentTicketUUID sutunu hic yoksa hata FIRLATMAZ",
    len(bo_parsed) == 1,
    bo_parsed,
)
check(
    "parse_excel_rows (BO/YÇM): parentTicketUUID sutunu yoksa bos string donuyor -> standalone ana ticket'a isaret eder",
    bo_parsed[0]["parent_ticket_uuid"] == "",
    bo_parsed[0],
)
check(
    "parse_excel_rows (BO/YÇM): tekrar eden 'Alternatif' basliklarinin HEPSI birlestiriliyor",
    bo_parsed[0]["alternative"] == "Ege Akdeniz Turu / 5 Gece Otel / 12.07 hareketli",
    bo_parsed[0]["alternative"],
)

otel_oda_excel = make_excel(
    [("111", "Otel Kaynaklı", "Alt 1", "", "Örnek Otel", "Double")],
    headers=("Rezervasyon No", "Kaydırma Tipi", "Alternatif 1", "parentTicketUUID", "Otel", "Oda Tipi"),
)
otel_oda_parsed = parse_excel_rows(otel_oda_excel)
check(
    "parse_excel_rows: opsiyonel Otel/Oda Tipi sutunlari okunuyor",
    otel_oda_parsed[0]["hotel"] == "Örnek Otel" and otel_oda_parsed[0]["room_type"] == "Double",
    otel_oda_parsed[0],
)

row_index_excel = make_excel(
    [
        ("111", "Otel Kaynaklı", "x", ""),
        (None, None, None, None),  # bos satir (atlanmali) -- sonraki satirin gercek Excel konumunu kaydirir
        ("222", "Operasyon Kaynaklı", "y", ""),
    ]
)
row_index_parsed = parse_excel_rows(row_index_excel)
check(
    "parse_excel_rows: excel_row_index bos satirlar atlansa bile GERCEK satir numarasini tasir (yazma icin gerekli)",
    row_index_parsed[0]["excel_row_index"] == 2 and row_index_parsed[1]["excel_row_index"] == 4,
    row_index_parsed,
)

# ==========================================================
# find_result_columns / generate_result_workbook (BO/YÇM sonuc dosyasi)
# ==========================================================
result_template_excel = make_excel(
    [
        ("346845477", "Operasyon Kaynaklı", "Alt", "", "", ""),
        # Gercek ornekte oldugu gibi AYNI rezervasyon no'nun birden fazla
        # satirda tekrar etmesi -- eslestirme reservation_no ile DEGIL,
        # excel_row_index (gercek satir numarasi) ile yapilmali.
        ("353230473", "Operasyon Kaynaklı", "Alt", "", "", ""),
        ("353230473", "Operasyon Kaynaklı", "Alt", "", "", ""),
    ],
    headers=("Rezervasyon No", "Kaydırma Tipi", "Alternatif 1", "Yeni Ticket ID", "Servis Durumu", "Servis Mesajı"),
)
result_columns = find_result_columns(result_template_excel)
check(
    "find_result_columns: BO/YÇM sonuc sutunlarinin 3'u de bulunuyor",
    result_columns is not None and set(result_columns.keys()) == {"ticket_id", "status", "message"},
    result_columns,
)

no_result_columns_excel = make_excel([("111", "Otel Kaynaklı", "x", "")])
check(
    "find_result_columns: sonuc sutunlari olmayan (Eos/wtatil) sablonda None doner",
    find_result_columns(no_result_columns_excel) is None,
)

result_template_excel.seek(0)
result_rows = parse_excel_rows(result_template_excel)
result_template_excel.seek(0)
original_bytes = result_template_excel.read()
fake_results = [
    {"ticket_id": "101946102", "success": True, "error": None},
    {"ticket_id": None, "success": False, "error": "HTTP 500"},
    {"ticket_id": "101946104", "success": True, "error": None},
]
generated_bytes = generate_result_workbook(original_bytes, result_rows, fake_results)
check("generate_result_workbook: BO/YÇM sablonunda bytes doner (None degil)", generated_bytes is not None)

from openpyxl import load_workbook as _load_workbook_for_test
import io as _io_for_test

readback_sheet = _load_workbook_for_test(_io_for_test.BytesIO(generated_bytes)).active
check(
    "generate_result_workbook: 1. satir (basarili) dogru satira yazilmis",
    readback_sheet.cell(row=2, column=4).value == "101946102"
    and readback_sheet.cell(row=2, column=5).value == "Başarılı",
    [readback_sheet.cell(row=2, column=c).value for c in range(4, 7)],
)
check(
    "generate_result_workbook: TEKRAR EDEN rezervasyon no'lu 2. ve 3. satirlar KARISMADAN, kendi satirina yaziliyor",
    readback_sheet.cell(row=3, column=4).value is None
    and readback_sheet.cell(row=3, column=5).value == "Hatalı"
    and readback_sheet.cell(row=3, column=6).value == "HTTP 500"
    and readback_sheet.cell(row=4, column=4).value == "101946104"
    and readback_sheet.cell(row=4, column=5).value == "Başarılı",
    [[readback_sheet.cell(row=r, column=c).value for c in range(4, 7)] for r in (3, 4)],
)

no_result_columns_excel.seek(0)
plain_rows = parse_excel_rows(no_result_columns_excel)
no_result_columns_excel.seek(0)
plain_bytes = no_result_columns_excel.read()
check(
    "generate_result_workbook: sonuc sutunu olmayan sablonda None doner (Eos/wtatil icin ek dosya YOK)",
    generate_result_workbook(plain_bytes, plain_rows, [{"ticket_id": "1", "success": True, "error": None}]) is None,
)

too_many_excel = make_excel([(str(i), "Otel Kaynaklı", "x", "uuid-1") for i in range(MAX_ROWS + 5)])
try:
    parse_excel_rows(too_many_excel)
    check(f"parse_excel_rows: {MAX_ROWS} satir sinirini asinca ValueError firlatir", False, "hata firlatmadi")
except ValueError as e:
    check(f"parse_excel_rows: {MAX_ROWS} satir sinirini asinca ValueError firlatir", str(MAX_ROWS) in str(e), str(e))

# ==========================================================
# Endpoint dogrulama (Flask test client) -- gercek CSM cagrisi yapilmayan yollar
# ==========================================================
client = app.test_client()
csrf_token = login_test_client(app, client)
csrf_headers = {"X-CSRF-Token": csrf_token}

resp = client.get("/api/bulk-shift/env")
check("GET /api/bulk-shift/env -> 200 + environment alani", resp.status_code == 200 and "environment" in resp.get_json())

resp = client.post("/api/bulk-shift/upload", data={}, headers=csrf_headers)
check("POST upload: dosya yoksa 400", resp.status_code == 400 and "dosya" in resp.get_json().get("error", "").lower())

# Not: parentTicketUUID eksikliği artık burada test EDİLMİYOR -- bir hata
# olmaktan çıktı (BO/YÇM'nin standalone "ana ticket" senaryosu, kullanıcı
# tarafından istendi), bu da o isteğin process_rows -> get_bulk_shift_token
# ile GERÇEK bir CSM ağ çağrısına kadar ilerlemesi anlamına gelir -- bu
# dosyanın kasıtlı olarak asla yapmadığı şey. O davranış yukarıda
# build_bulk_ticket_payload/parse_excel_rows seviyesinde (ağsız) test edildi.

resp = client.post(
    "/api/bulk-shift/upload",
    data={
        "file": (make_excel([("x", "y", "z", "w")], headers=("Yanlis", "Sutunlar", "Burada", "Da")), "bad.xlsx"),
        "parent_ticket_uuid": "abc",
    },
    content_type="multipart/form-data",
    headers=csrf_headers,
)
check(
    "POST upload: yanlis sutunlu Excel icin 400 (Rezervasyon No bulunamadi)",
    resp.status_code == 400 and "Rezervasyon No" in resp.get_json().get("error", ""),
    resp.get_json(),
)

# ==========================================================
# POST upload: BASARILI (process_rows mock'lanarak, gercek CSM cagrisi
# olmadan) sonuc mail_processing_log.jsonl'a yaziliyor mu -- CANLI HATA
# DUZELTMESI: bu satir hic yoktu, panelden yapilan toplu kaydirmalar
# (basarili VEYA basarisiz) Loglar sayfasinda hic gorunmuyordu.
# ==========================================================
_original_process_rows = app_module.process_rows


def _fake_process_rows(rows, fallback_parent_ticket_uuid=""):
    return {
        "environment": "prod",
        "total": 2,
        "success_count": 0,
        "failed_count": 2,
        "results": [
            {
                "reservation_no": r["reservation_no"], "shift_type": r["shift_type"],
                "shift_type_code": "OPERASYON_KAYNAKLI", "is_linked": bool(r["parent_ticket_uuid"]),
                "success": False, "ticket_id": None, "error": "HTTP 500: Ticket kapatilamaz.",
            }
            for r in rows
        ],
    }


app_module.process_rows = _fake_process_rows
try:
    resp = client.post(
        "/api/bulk-shift/upload",
        data={"file": (make_excel([("111", "Otel Kaynaklı", "x", ""), ("222", "Otel Kaynaklı", "y", "")]), "test.xlsx")},
        content_type="multipart/form-data",
        headers=csrf_headers,
    )
finally:
    app_module.process_rows = _original_process_rows

check("POST upload (mock): basarisiz sonuc bile 200 donuyor (ozet JSON'da)", resp.status_code == 200, resp.get_json())

mail_logs = logging_utils.read_mail_events(limit=10)
check(
    "POST upload (CANLI HATA DUZELTMESI): panel yuklemesi artik mail_processing_log'a yaziliyor",
    len(mail_logs) == 1 and mail_logs[0]["classification"] == "bulk_kaydirma",
    mail_logs,
)
check(
    "POST upload: tum satirlar basarisizsa status='failed' olarak logluyor",
    len(mail_logs) == 1 and mail_logs[0]["status"] == "failed" and mail_logs[0]["event"] == "ticket_not_created",
    mail_logs[0] if mail_logs else None,
)
check(
    "POST upload: kim yukledigi (panel kullanicisi) sender_email olarak kaydediliyor",
    len(mail_logs) == 1 and mail_logs[0]["sender_email"] == "tester@local.test",
    mail_logs[0] if mail_logs else None,
)

if os.path.exists(_tmp_mail_log):
    os.remove(_tmp_mail_log)

# ==========================================================
# OZET
# ==========================================================
print("=" * 60)
for status, name, detail in results:
    mark = "[OK]" if status == "PASS" else "[FAIL]"
    line = f"{mark} {name}"
    if status == "FAIL":
        line += f"  -> {detail}"
    print(line)

print("=" * 60)
print(f"TOPLAM: {passed + failed} senaryo | BASARILI: {passed} | BASARISIZ: {failed}")
if failed:
    raise SystemExit(1)
