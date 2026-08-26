"""Toplu Kaydırma (Bulk Reservation Shift) module.

Creates CSM tickets in bulk, one per reservation row, from an uploaded Excel
export of reservations that need a date shift. Two iş birimi usage patterns,
both driven by a single per-row rule -- NOT a role/team toggle:
  - Eos / wtatil: every row's parentTicketUUID column is filled in -> each
    row becomes an "ilişkili ticket" (LINKED_TICKET) hanging off that main
    ticket.
  - BO / YÇM: parentTicketUUID is left blank (their export doesn't even have
    that column) -> each row becomes its own standalone, independent main
    ticket ("ana ticket") -- no relation, no parent.
A single upload can freely mix both (row-by-row), see build_bulk_ticket_payload.

Payload shape verified against a real, successful production request
(captured from the CSM web UI's Network tab, 2026-08-25, ticket #101944819,
HTTP 200) -- see build_bulk_ticket_payload for the exact fields. Two things
about that captured example that are NOT obvious from CSM's UI:
  1. The link to the main ticket is two top-level fields on the payload
     (parentTicketUUID + relationType: "LINKED_TICKET") -- NOT an entry in
     ticketRelationList (which stays empty, unlike the parent-child relation
     this module used before). For a standalone "ana ticket" both fields are
     simply omitted -- same kırılım/reporter otherwise (kullanıcı tarafından
     onaylandı, 2026-08-26), only the relation is different.
  2. The reporter/"raporlayan kişi" for every ticket is a fixed, pre-existing
     anonymous CSM contact ("Onay Kaydırma", onay@tatilbudur.com) -- not the
     individual mail sender, and not something collected via a form.
"""

import html
import os
from typing import Any, BinaryIO, Dict, List, Optional, TypedDict

import requests
from openpyxl import load_workbook

from service_log import record_service_event

MAX_ROWS = 500
MAX_EXCEL_BYTES = 5 * 1024 * 1024  # 500 satırlık gerçek bir dosya birkaç yüz KB'dir

# Required columns -- exact match from the source Excel export (Eos/wtatil
# AND BO/YÇM templates both use these same three names).
REQUIRED_COLUMN_ALIASES = {
    "reservation_no": ("Rezervasyon No",),
    "shift_type": ("Kaydırma Tipi",),
}
# Optional columns -- absent entirely in some templates (ör. BO/YÇM'nin
# parentTicketUUID sütunu hiç yok), tolerated as empty/"" per row.
OPTIONAL_COLUMN_ALIASES = {
    "parent_ticket_uuid": ("parentTicketUUID",),
    "hotel": ("Otel",),
    "room_type": ("Oda Tipi",),
}
# Result columns some templates (BO/YÇM) expect filled back in and returned
# as a downloadable file -- see generate_result_workbook.
RESULT_COLUMN_ALIASES = {
    "ticket_id": ("Yeni Ticket ID",),
    "status": ("Servis Durumu",),
    "message": ("Servis Mesajı",),
}
# "Alternatif 1"/"Alternatif 2"/"Alternatif 3", or several bare "Alternatif"
# columns (both observed live) -- ANY header starting with this (case-
# insensitive) is collected, in column order, and joined for the description.
ALTERNATIVE_HEADER_PREFIX = "alternatif"


class ExcelRow(TypedDict):
    reservation_no: str
    shift_type: str
    alternative: str
    parent_ticket_uuid: str
    hotel: str
    room_type: str
    excel_row_index: int


def _normalized_headers(header_row) -> List[str]:
    return [str(cell).strip() if cell is not None else "" for cell in header_row]


def parse_excel_rows(file_stream: BinaryIO) -> List[ExcelRow]:
    """Read the uploaded Excel export into plain dicts.

    Raises ValueError if a required column is missing or the sheet has more
    than MAX_ROWS data rows (a hard cap — this endpoint processes rows
    synchronously, one HTTP request per row to CSM, so an unbounded upload
    would tie up a request/worker for an unbounded amount of time). Also
    raises ValueError if the raw file exceeds MAX_EXCEL_BYTES -- both mail
    attachments and panel uploads are untrusted input, and openpyxl parsing
    an oversized/crafted .xlsx (zip decompression bomb) before MAX_ROWS
    ever gets a chance to apply is a DoS vector.
    """
    file_stream.seek(0, os.SEEK_END)
    size = file_stream.tell()
    file_stream.seek(0)
    if size > MAX_EXCEL_BYTES:
        raise ValueError(
            f"Excel dosyası çok büyük ({size // 1024} KB) — üst sınır {MAX_EXCEL_BYTES // 1024} KB."
        )

    workbook = load_workbook(filename=file_stream, data_only=True, read_only=True)
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ValueError("Excel dosyası boş görünüyor.")

    headers = _normalized_headers(header_row)

    column_positions: Dict[str, int] = {}
    for field, aliases in REQUIRED_COLUMN_ALIASES.items():
        idx = next((i for i, h in enumerate(headers) if h in aliases), None)
        if idx is None:
            raise ValueError(f"Beklenen sütun bulunamadı: '{aliases[0]}'")
        column_positions[field] = idx

    for field, aliases in OPTIONAL_COLUMN_ALIASES.items():
        idx = next((i for i, h in enumerate(headers) if h in aliases), None)
        if idx is not None:
            column_positions[field] = idx

    alternative_indices = [
        i for i, h in enumerate(headers) if h.lower().startswith(ALTERNATIVE_HEADER_PREFIX)
    ]

    rows: List[ExcelRow] = []
    last_parent_ticket_uuid = ""
    for excel_row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None or all(cell is None for cell in row):
            continue
        if len(rows) >= MAX_ROWS:
            raise ValueError(f"Excel dosyası {MAX_ROWS} satırdan fazla içeriyor — daha küçük parçalara bölün.")

        def cell(field: str) -> str:
            pos = column_positions.get(field)
            if pos is None:
                return ""
            value = row[pos] if pos < len(row) else None
            return "" if value is None else str(value).strip()

        reservation_no = cell("reservation_no")
        if not reservation_no:
            continue

        alternative_parts = []
        for idx in alternative_indices:
            value = row[idx] if idx < len(row) else None
            if value is not None and str(value).strip():
                alternative_parts.append(str(value).strip())

        # parentTicketUUID is typically only filled on the first row of each
        # WorkGroup/Ticket ID group in the source export (observed live) --
        # carry the last non-empty value forward for rows that leave it blank.
        # Templates with no parentTicketUUID column at all (BO/YÇM) simply
        # never populate this -- every row's parent_ticket_uuid stays "",
        # which build_bulk_ticket_payload treats as "standalone ana ticket".
        parent_ticket_uuid = cell("parent_ticket_uuid") or last_parent_ticket_uuid
        if parent_ticket_uuid:
            last_parent_ticket_uuid = parent_ticket_uuid

        rows.append(
            ExcelRow(
                reservation_no=reservation_no,
                shift_type=cell("shift_type") or "Operasyon Kaynaklı",
                alternative=" / ".join(alternative_parts) or "Toplu Kaydırma İşlemi",
                parent_ticket_uuid=parent_ticket_uuid,
                hotel=cell("hotel"),
                room_type=cell("room_type"),
                excel_row_index=excel_row_index,
            )
        )

    return rows


def find_result_columns(file_stream: BinaryIO) -> Optional[Dict[str, int]]:
    """Whether the uploaded template has the BO/YÇM result columns ("Yeni
    Ticket ID" / "Servis Durumu" / "Servis Mesajı") -- if all three are
    present, generate_result_workbook can fill them in and the caller should
    offer the filled-in file back as a download/attachment. Returns None
    (nothing to fill in) for templates missing any of the three, e.g. the
    plain Eos/wtatil template."""
    file_stream.seek(0)
    workbook = load_workbook(filename=file_stream, data_only=True, read_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    file_stream.seek(0)
    if not header_row:
        return None

    headers = _normalized_headers(header_row)
    positions: Dict[str, int] = {}
    for field, aliases in RESULT_COLUMN_ALIASES.items():
        idx = next((i for i, h in enumerate(headers) if h in aliases), None)
        if idx is None:
            return None
        positions[field] = idx
    return positions


def generate_result_workbook(
    original_bytes: bytes, rows: List[ExcelRow], results: List[Dict[str, Any]]
) -> Optional[bytes]:
    """Writes each row's outcome back into the ORIGINAL uploaded file's "Yeni
    Ticket ID" / "Servis Durumu" / "Servis Mesajı" columns (BO/YÇM template)
    and returns the updated file's bytes -- or None if that template's
    columns aren't present (nothing to fill in, e.g. Eos/wtatil's simpler
    template). Matches rows by their exact worksheet row number
    (ExcelRow.excel_row_index), NOT by reservation_no -- the same
    reservation number can legitimately repeat across multiple rows
    (observed live, different room/shift entries under one reservation)."""
    import io

    result_columns = find_result_columns(io.BytesIO(original_bytes))
    if result_columns is None:
        return None

    workbook = load_workbook(filename=io.BytesIO(original_bytes))
    sheet = workbook.active

    for row, outcome in zip(rows, results):
        excel_row = row["excel_row_index"]
        sheet.cell(row=excel_row, column=result_columns["ticket_id"] + 1).value = outcome.get("ticket_id") or ""
        sheet.cell(row=excel_row, column=result_columns["status"] + 1).value = (
            "Başarılı" if outcome.get("success") else "Hatalı"
        )
        sheet.cell(row=excel_row, column=result_columns["message"] + 1).value = outcome.get("error") or ""

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()

BULK_SHIFT_ENV = os.getenv("BULK_SHIFT_ENV", "preprod")

_ENV_URLS = {
    "preprod": {
        "auth": "https://tatilbudur-pp-api.cloudcsmetiya.com/api/v1/auth/authenticate",
        "create_ticket": "https://tatilbudur-pp-api.cloudcsmetiya.com/api/v1/business/ticket/add",
    },
    "prod": {
        "auth": "https://tatilbudur-api.cloudcsmetiya.com/api/v1/auth/authenticate",
        "create_ticket": "https://tatilbudur-api.cloudcsmetiya.com/api/v1/business/ticket/add",
    },
}

# The fixed "Onay Kaydırma" anonymous contact every linked ticket is reported
# by (verified real party role id -- reusing it keeps CSM from accumulating
# more near-duplicate "onay kaydırma"-ish contacts, several of which already
# exist from past manual use).
ONAY_KAYDIRMA_PARTY_ROLE = {
    "id": 100239153,
    "party": {
        "firstName": "Onay",
        "lastName": "Kaydırma",
        "nationalityId": None,
        "partyType": "INDIVIDUAL",
        "fullName": "Onay Kaydırma",
    },
    "partyRoleType": {
        "id": 1000022,
        "uuid": "4b7fc8ce-289c-476f-92d0-9279e7199983",
        "shortCode": "ANONYMOUS",
        "ownerType": True,
    },
    "externalId": None,
    "attributeValueList": [],
    "contactMediumList": [
        {
            "contactMediumType": {"id": 10000005, "uuid": "c38936ac-7e2c-46e7-aacc-deca59e132bf", "shortCode": "EMAIL"},
            "val": "onay@tatilbudur.com",
            "isPrimary": True,
            "externalId": None,
        }
    ],
}

# Channel/ticketType/category/subCategory per "Kaydırma Tipi" classification
# (kullanıcı tarafından doğrulandı, 2026-08-25). The relation to the main
# ticket (top-level parentTicketUUID + relationType, see build_bulk_ticket_payload)
# and the fixed "Onay Kaydırma" reporter are shared across all three; only the
# kırılım (category path) differs by shift type. Add more entries here as new
# kırılımlar are confirmed.
SHIFT_TYPE_FIELDS: Dict[str, Dict[str, Any]] = {
    "OTEL_KAYNAKLI": {
        "channel": {"shortCode": "CAGRI_MERKEZI", "name": "Çağrı Merkezi", "id": 100000041, "uuid": "9a4dd8ef-045a-4f57-8c2c-8b0fc209d367"},
        "ticketType": {"shortCode": "REZERVASYON_ISLEMLERI", "name": "Backoffice İşlemleri", "id": 100000059, "uuid": "654c901c-40c2-46c1-8a00-925009dde46e"},
        "category": {"shortCode": "KAYDIRMA", "name": "Kaydırma", "id": 100000143, "uuid": "f3a9a508-0732-4410-82d7-847867fc616f"},
        "subCategory": {"shortCode": "OTEL_KAYNAKLI", "name": "Otel Kaynaklı", "id": 100000669, "uuid": "089de3d8-b26d-42da-88e6-f062cae187f3"},
    },
    "OPERASYON_KAYNAKLI": {
        "channel": {"shortCode": "CAGRI_MERKEZI", "name": "Çağrı Merkezi", "id": 100000041, "uuid": "9a4dd8ef-045a-4f57-8c2c-8b0fc209d367"},
        "ticketType": {"shortCode": "REZERVASYON_ISLEMLERI", "name": "Backoffice İşlemleri", "id": 100000059, "uuid": "654c901c-40c2-46c1-8a00-925009dde46e"},
        "category": {"shortCode": "KAYDIRMA", "name": "Kaydırma", "id": 100000143, "uuid": "f3a9a508-0732-4410-82d7-847867fc616f"},
        "subCategory": {"shortCode": "OPERASYON_KAYNAKLI", "name": "Operasyon Kaynaklı", "id": 100000671, "uuid": "089de3d8-b26d-42da-88e6-f062cae187f3"},
    },
    "ODEME_TAMAMLAMA": {
        "channel": {"shortCode": "CAGRI_MERKEZI", "name": "Çağrı Merkezi", "id": 100000041, "uuid": "9a4dd8ef-045a-4f57-8c2c-8b0fc209d367"},
        "ticketType": {"shortCode": "REZERVASYON_ISLEMLERI", "name": "Backoffice İşlemleri", "id": 100000059, "uuid": "654c901c-40c2-46c1-8a00-925009dde46e"},
        "category": {"shortCode": "DIGER_ISLEMLER", "name": "Diğer İşlemler", "id": 100000172, "uuid": "f3a9a508-0732-4410-82d7-847867fc616f"},
        "subCategory": {"shortCode": "ODEME_TAMAMLAMA", "name": "Ödeme Tamamlama", "id": 100000554, "uuid": "089de3d8-b26d-42da-88e6-f062cae187f3"},
    },
}

_cached_token: Optional[str] = None


def classify_shift_type(kaydirma_tipi: str) -> str:
    """Mirrors the original script's keyword-based classification -- selects
    which SHIFT_TYPE_FIELDS entry (channel/category/subCategory/ticketType)
    a row's ticket uses."""
    text = (kaydirma_tipi or "").strip()
    if "Otel" in text:
        return "OTEL_KAYNAKLI"
    if "Ödeme" in text or "Tamamlama" in text:
        return "ODEME_TAMAMLAMA"
    return "OPERASYON_KAYNAKLI"


def get_bulk_shift_token(force_refresh: bool = False) -> str:
    """Get a bearer token for the configured environment (BULK_SHIFT_ENV).

    Uses the dedicated RPA_OTOMASYON_USERNAME/PASSWORD system account (falls
    back to CSM_USERNAME/CSM_PASSWORD if that account isn't provisioned yet)
    so linked tickets show "Oluşturan: rpa_otomasyon" in CSM rather than a
    personal account.
    """
    global _cached_token
    if _cached_token and not force_refresh:
        return _cached_token

    from config import RPA_OTOMASYON_USERNAME, RPA_OTOMASYON_PASSWORD

    if not RPA_OTOMASYON_USERNAME or not RPA_OTOMASYON_PASSWORD:
        raise RuntimeError("RPA_OTOMASYON_USERNAME / RPA_OTOMASYON_PASSWORD (veya CSM_USERNAME / CSM_PASSWORD) ayarlanmamış.")

    auth_url = _ENV_URLS[BULK_SHIFT_ENV]["auth"]
    try:
        response = requests.post(
            auth_url,
            json={"userName": RPA_OTOMASYON_USERNAME, "password": RPA_OTOMASYON_PASSWORD},
            headers={"Content-Type": "application/json", "Accept": "application/json"},
        )
        if response.status_code != 200:
            record_service_event(
                "csm_api", "auth", "failed",
                detail=f"[{BULK_SHIFT_ENV.upper()}] HTTP {response.status_code}",
            )
            raise RuntimeError(f"Token alınamadı. Status Code: {response.status_code}")

        auth_header = response.headers.get("authorization") or response.headers.get("Authorization") or ""
        if auth_header:
            token = auth_header.replace("Bearer ", "").replace("bearer ", "").strip()
        else:
            res_data = response.json()
            token = res_data.get("token") or res_data.get("access_token")

        if not token:
            record_service_event(
                "csm_api", "auth", "failed",
                detail=f"[{BULK_SHIFT_ENV.upper()}] Token yanıtta bulunamadı",
            )
            raise RuntimeError("Token yanıtta bulunamadı.")

        _cached_token = token
        record_service_event("csm_api", "auth", "success", detail=f"[{BULK_SHIFT_ENV.upper()}] Token alındı")
        return token
    except requests.RequestException as e:
        record_service_event("csm_api", "auth", "failed", detail=f"[{BULK_SHIFT_ENV.upper()}] {e}")
        raise RuntimeError(str(e))


def build_bulk_ticket_payload(
    reservation_no: str,
    shift_type_label: str,
    alternative_text: str,
    parent_ticket_uuid: str,
    hotel: str = "",
    room_type: str = "",
) -> Dict[str, Any]:
    """parent_ticket_uuid empty ("") -> standalone "ana ticket" (BO/YÇM):
    parentTicketUUID/relationType are omitted entirely, nothing to link to.
    parent_ticket_uuid set -> "ilişkili ticket" (Eos/wtatil), same as before.
    Kırılım (channel/category/subCategory/ticketType) and reporter are
    IDENTICAL either way (kullanıcı tarafından onaylandı, 2026-08-26) --
    only the relation to a parent differs."""
    party_role = ONAY_KAYDIRMA_PARTY_ROLE
    fields = SHIFT_TYPE_FIELDS[classify_shift_type(shift_type_label)]

    description_parts = [
        f"Toplu Kaydırma - Rezervasyon: {html.escape(str(reservation_no))} - "
        f"{html.escape(shift_type_label)} - {html.escape(alternative_text)}"
    ]
    if hotel:
        description_parts.append(f"Otel: {html.escape(hotel)}")
    if room_type:
        description_parts.append(f"Oda Tipi: {html.escape(room_type)}")

    payload: Dict[str, Any] = {
        "contactParties": [
            {
                "partyRole": party_role,
                "contactMediumList": [
                    {"contactMedium": party_role["contactMediumList"][0], "isPreferred": True},
                ],
                "isPreferred": True,
                "preferredContactChannelList": ["EMAIL"],
            }
        ],
        "attributeList": [],
        "attributeSetList": [],
        "ticketRelationList": [],
        "subTicketList": [],
        "availableOperations": [],
        "stage": {"shortCode": "START"},
        "partyRole": party_role,
        # Excel'den gelen değerler (mail ekinden -- güvenilmeyen kaynak da
        # olabilir) HTML açıklama içine gömülmeden önce escape edilmeli;
        # CSM ticket açıklamasını kendi arayüzünde HTML olarak render ediyor.
        "description": "<p>" + " - ".join(description_parts) + "</p>",
        "channel": fields["channel"],
        "subCategory": fields["subCategory"],
        "category": fields["category"],
        "ticketType": fields["ticketType"],
        "priorityLevel": {
            "shortCode": "NORMAL",
            "name": "Normal",
            "ordinal": 3,
            "createdInAI": None,
            "externalId": None,
            "expression": None,
            "extraFields": {"color": "#FFE733"},
            "id": 100000037,
            "uuid": "074266b0-efdf-45a1-ab26-75a627a4b5ed",
        },
        "isResolved": False,
        "isProactive": False,
        "isParent": True,
        "reminders": [],
        "detectedLanguage": "",
        "relatedProduct": {"serviceNumber": str(reservation_no)},
    }
    if parent_ticket_uuid:
        payload["parentTicketUUID"] = parent_ticket_uuid
        payload["relationType"] = "LINKED_TICKET"
    return payload


def create_bulk_ticket(payload: Dict[str, Any], token: str) -> Dict[str, Any]:
    """POST one bulk-shift ticket. Returns {'success', 'ticket_id'?, 'error'?}."""
    import json

    url = _ENV_URLS[BULK_SHIFT_ENV]["create_ticket"]
    headers = {
        "Authorization": f"Bearer {token}",
        "Multilanguage": "true",
        "userMode": "LEAD",
        "Accept": "application/json, text/plain, */*",
    }
    try:
        payload_json = json.dumps(payload, ensure_ascii=False)
        files = {"ticket": ("blob", payload_json.encode("utf-8"), "application/json")}
        response = requests.post(url, files=files, headers=headers, timeout=30)

        if response.status_code in (200, 201):
            res_data = response.json()
            ticket_id = res_data.get("id") or res_data.get("ticketNo") or "Oluşturuldu"
            record_service_event(
                "csm_api", "create_bulk_ticket", "success",
                detail=f"[{BULK_SHIFT_ENV.upper()}] #{ticket_id}",
            )
            return {"success": True, "ticket_id": str(ticket_id)}

        record_service_event(
            "csm_api", "create_bulk_ticket", "failed",
            detail=f"[{BULK_SHIFT_ENV.upper()}] HTTP {response.status_code}",
        )
        return {"success": False, "error": f"HTTP {response.status_code}: {response.text[:200]}"}
    except requests.RequestException as e:
        record_service_event("csm_api", "create_bulk_ticket", "failed", detail=f"[{BULK_SHIFT_ENV.upper()}] {e}")
        return {"success": False, "error": str(e)}


def process_rows(rows: List[ExcelRow], fallback_parent_ticket_uuid: str = "") -> Dict[str, Any]:
    """Resolves a bearer token once, then creates one linked ticket per row.
    Shared by the panel's /api/bulk-shift/upload endpoint and the
    mail-triggered flow (main.py's process_email, "toplu kaydırma" subject +
    Excel attachment) so both go through the exact same CSM call logic."""
    token = get_bulk_shift_token()

    results = []
    success_count = 0
    for row in rows:
        parent_ticket_uuid = row["parent_ticket_uuid"] or fallback_parent_ticket_uuid
        payload = build_bulk_ticket_payload(
            reservation_no=row["reservation_no"],
            shift_type_label=row["shift_type"],
            alternative_text=row["alternative"],
            parent_ticket_uuid=parent_ticket_uuid,
            hotel=row.get("hotel", ""),
            room_type=row.get("room_type", ""),
        )
        outcome = create_bulk_ticket(payload, token)
        if outcome["success"]:
            success_count += 1
        results.append({
            "reservation_no": row["reservation_no"],
            "shift_type": row["shift_type"],
            "shift_type_code": classify_shift_type(row["shift_type"]),
            "is_linked": bool(parent_ticket_uuid),
            "success": outcome["success"],
            "ticket_id": outcome.get("ticket_id"),
            "error": outcome.get("error"),
        })

    return {
        "environment": BULK_SHIFT_ENV,
        "total": len(results),
        "success_count": success_count,
        "failed_count": len(results) - success_count,
        "results": results,
    }
